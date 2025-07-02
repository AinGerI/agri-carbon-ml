#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PCA Analyzer for Agricultural Data

This module provides comprehensive Principal Component Analysis capabilities including:
- Singular value decomposition analysis for each province/region
- Variance contribution rate calculation and visualization
- Iterative analysis with random indicator removal
- Multi-threshold analysis (80%, 85%, 90%, 95% variance)
- Rich visualizations and comprehensive reporting

Combines functionality from:
- 主成分分析确定贡献值.py 
- 迭代寻找最多的主成分.py

Author: Thesis Research Project
Date: 2025
"""

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.linalg import svd
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import time
import random
import io
import shutil
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

# Optional rich library for enhanced console output
try:
    from rich.progress import Progress, TextColumn, BarColumn, TimeElapsedColumn, TimeRemainingColumn
    from rich.console import Console
    from rich.panel import Panel
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False

# Configure matplotlib for Chinese font support
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'SimSun', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False
plt.rcParams['font.family'] = 'sans-serif'

# Attempt to load additional Chinese font support
try:
    import matplotlib as mpl
    mpl.font_manager.fontManager.addfont('SimHei.ttf')
except:
    pass

if RICH_AVAILABLE:
    console = Console()


class PCAAnalyzer:
    """
    Comprehensive PCA analyzer for agricultural research data.
    
    Supports both single analysis and iterative analysis with random indicator removal.
    """
    
    def __init__(self):
        """Initialize with default analysis parameters."""
        self.variance_thresholds = [0.8, 0.85, 0.9, 0.95]
        self.threshold_colors = ['#ff7f0e', '#d62728', '#9467bd', '#8c564b']
        self.default_remove_percentage = 0.5
        self.default_iterations = 20
        
    def _log_message(self, message, use_rich=True):
        """Log message with optional rich formatting."""
        if RICH_AVAILABLE and use_rich:
            console.print(message)
        else:
            print(message)
    
    def _create_output_dirs(self, base_dir, iteration=None):
        """Create organized output directory structure."""
        if iteration is not None:
            iteration_dir = os.path.join(base_dir, f'迭代{iteration}')
            dirs = {
                'root': iteration_dir,
                'single_plots': os.path.join(iteration_dir, '单省市奇异值图'),
                'summary_plots': os.path.join(iteration_dir, '汇总图表'),
                'data_tables': os.path.join(iteration_dir, '数据表格'),
                'temp': os.path.join(iteration_dir, '临时文件')
            }
        else:
            dirs = {
                'root': base_dir,
                'single_plots': os.path.join(base_dir, '单省市奇异值图'),
                'summary_plots': os.path.join(base_dir, '汇总图表'),
                'data_tables': os.path.join(base_dir, '数据表格'),
                'temp': os.path.join(base_dir, '临时文件')
            }
        
        # Create directories
        for dir_path in dirs.values():
            os.makedirs(dir_path, exist_ok=True)
        
        return dirs
    
    def _calculate_principal_components_needed(self, singular_values_dict):
        """Calculate required principal components for different variance thresholds."""
        results = []
        
        for sheet_name, s in singular_values_dict.items():
            # Calculate variance contribution rates
            explained_variance_ratio = s**2 / np.sum(s**2)
            cumulative_explained_variance = np.cumsum(explained_variance_ratio)
            
            # Find required components for each threshold
            components_needed = {}
            for threshold in self.variance_thresholds:
                idx = np.where(cumulative_explained_variance >= threshold)[0]
                if len(idx) > 0:
                    components_needed[threshold] = idx[0] + 1
                else:
                    components_needed[threshold] = len(s)
            
            results.append({
                '省市': sheet_name,
                '总维度': len(s),
                '奇异值': s.tolist(),
                '方差贡献率': explained_variance_ratio.tolist(),
                '累积方差贡献率': cumulative_explained_variance.tolist(),
                '80%方差所需主成分': components_needed[0.8],
                '85%方差所需主成分': components_needed[0.85],
                '90%方差所需主成分': components_needed[0.9],
                '95%方差所需主成分': components_needed[0.95],
            })
        
        return results
    
    def _create_province_plot(self, sheet_name, s, explained_variance_ratio, output_dir):
        """Create individual province analysis plots."""
        cumulative_explained_variance = np.cumsum(explained_variance_ratio)
        
        # Create figure with two subplots
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
        
        # Subplot 1: Singular values
        ax1.plot(range(1, len(s) + 1), s, 'o-', markersize=8, color='#1f77b4')
        ax1.set_title(f'{sheet_name} - 奇异值分布', fontsize=14)
        ax1.set_xlabel('主成分索引', fontsize=12)
        ax1.set_ylabel('奇异值', fontsize=12)
        ax1.grid(True, alpha=0.5)
        
        # Subplot 2: Cumulative variance contribution
        ax2.plot(range(1, len(cumulative_explained_variance) + 1), 
                cumulative_explained_variance * 100, 'o-', markersize=8, color='#2ca02c')
        
        # Add threshold lines
        thresholds = [80, 85, 90, 95]
        for threshold, color in zip(thresholds, self.threshold_colors):
            ax2.axhline(y=threshold, color=color, linestyle='--', 
                       label=f'{threshold}%阈值', alpha=0.7)
            
            # Mark threshold intersections
            idx = np.where(cumulative_explained_variance * 100 >= threshold)[0]
            if len(idx) > 0:
                pc_count = idx[0] + 1
                ax2.plot(pc_count, threshold, 'o', color=color, markersize=8)
                ax2.annotate(f'PC{pc_count}', 
                           xy=(pc_count, threshold), 
                           xytext=(pc_count+0.5, threshold+2),
                           arrowprops=dict(facecolor=color, shrink=0.05, width=1.5, headwidth=8),
                           fontsize=10)
        
        ax2.set_title(f'{sheet_name} - 累积方差贡献率', fontsize=14)
        ax2.set_xlabel('主成分数量', fontsize=12)
        ax2.set_ylabel('累积方差贡献率 (%)', fontsize=12)
        ax2.grid(True, alpha=0.5)
        ax2.legend(loc='lower right')
        
        plt.tight_layout()
        
        # Save plot
        file_path = os.path.join(output_dir, f'{sheet_name}_奇异值分析.png')
        plt.savefig(file_path, dpi=300, bbox_inches='tight')
        plt.close(fig)
        
        return fig, file_path
    
    def _save_singular_values_to_excel(self, singular_values_dict, output_dir):
        """Save all singular values data to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "奇异值数据"
        
        # Write headers
        ws['A1'] = "省市"
        max_length = max(len(s) for s in singular_values_dict.values())
        
        for i in range(max_length):
            ws.cell(row=1, column=i+2).value = f"奇异值{i+1}"
        
        # Write data
        row = 2
        for province, values in singular_values_dict.items():
            ws.cell(row=row, column=1).value = province
            for i, val in enumerate(values):
                ws.cell(row=row, column=i+2).value = val
            row += 1
        
        excel_path = os.path.join(output_dir, '各省市奇异值.xlsx')
        wb.save(excel_path)
        return excel_path
    
    def _create_summary_excel(self, results, output_dirs):
        """Create comprehensive summary Excel file."""
        df = pd.DataFrame(results)
        df_sorted = df.sort_values(by='80%方差所需主成分')
        
        excel_path = os.path.join(output_dirs['data_tables'], '主成分分析汇总.xlsx')
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Main components summary
            df_sorted[['省市', '总维度', '80%方差所需主成分', '85%方差所需主成分', 
                       '90%方差所需主成分', '95%方差所需主成分']].to_excel(
                writer, sheet_name='主成分数量', index=False)
            
            # Singular values data
            singular_values_df = pd.DataFrame({'省市': df['省市']})
            max_length = max(len(s) for s in df['奇异值'])
            
            for i in range(max_length):
                singular_values_df[f'奇异值{i+1}'] = df['奇异值'].apply(
                    lambda x: x[i] if i < len(x) else None)
            
            singular_values_df.to_excel(writer, sheet_name='奇异值数据', index=False)
            
            # Variance contribution data
            variance_df = pd.DataFrame({'省市': df['省市']})
            
            for i in range(max_length):
                variance_df[f'PC{i+1}方差贡献率'] = df['方差贡献率'].apply(
                    lambda x: x[i] if i < len(x) else None)
                variance_df[f'PC{i+1}累积方差贡献率'] = df['累积方差贡献率'].apply(
                    lambda x: x[i] if i < len(x) else None)
            
            variance_df.to_excel(writer, sheet_name='方差贡献率', index=False)
        
        return excel_path
    
    def _randomly_remove_indicators(self, file_path, remove_percentage=0.5):
        """Randomly remove a percentage of indicators and save as new file."""
        temp_dir = os.path.join(os.path.dirname(file_path), 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        temp_file = os.path.join(temp_dir, f'随机删除指标_{timestamp}.xlsx')
        
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        removed_indicators = {}
        
        with pd.ExcelWriter(temp_file, engine='openpyxl') as writer:
            for sheet_name in sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name)
                
                if df.shape[0] > 0:
                    indicators = df.iloc[:, 0].tolist()
                    num_to_remove = int(len(indicators) * remove_percentage)
                    indices_to_remove = random.sample(range(len(indicators)), num_to_remove)
                    
                    removed_indicators[sheet_name] = [indicators[i] for i in indices_to_remove]
                    new_df = df.drop(indices_to_remove)
                    new_df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        return temp_file, removed_indicators
    
    def analyze_singular_values(self, file_path, output_dir=None, create_plots=True):
        """
        Perform comprehensive singular value analysis on Excel file.
        
        Args:
            file_path: Path to Excel file with province data
            output_dir: Output directory (auto-generated if None)
            create_plots: Whether to create visualization plots
            
        Returns:
            dict: Analysis results including paths and statistics
        """
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(file_path), '奇异值分析结果')
        
        output_dirs = self._create_output_dirs(output_dir)
        
        if RICH_AVAILABLE:
            console.print(Panel.fit(
                "[bold green]奇异值分析开始[/bold green]\n"
                f"[yellow]输入文件:[/yellow] {file_path}\n"
                f"[yellow]输出目录:[/yellow] {output_dir}"
            ))
        else:
            print(f"开始奇异值分析\n输入文件: {file_path}\n输出目录: {output_dir}")
        
        # Read Excel file
        excel_file = pd.ExcelFile(file_path)
        sheet_names = excel_file.sheet_names
        
        singular_values_dict = {}
        
        # Create summary plot for all provinces
        if create_plots:
            fig_all, ax_all = plt.subplots(figsize=(15, 10))
        
        # Progress tracking
        if RICH_AVAILABLE:
            with Progress(
                TextColumn("[bold blue]{task.description}"),
                BarColumn(bar_width=40),
                TextColumn("[bold green]{task.completed}/{task.total}"),
                TimeElapsedColumn(),
                TimeRemainingColumn(),
            ) as progress:
                main_task = progress.add_task("[red]处理省市数据...", total=len(sheet_names))
                
                for i, sheet_name in enumerate(sheet_names):
                    progress.update(main_task, description=f"[red]处理 {sheet_name}")
                    self._process_sheet(excel_file, sheet_name, singular_values_dict, output_dirs, create_plots, fig_all if create_plots else None)
                    progress.update(main_task, advance=1)
                    time.sleep(0.01)
        else:
            for i, sheet_name in enumerate(sheet_names):
                print(f"处理 {sheet_name} ({i+1}/{len(sheet_names)})")
                self._process_sheet(excel_file, sheet_name, singular_values_dict, output_dirs, create_plots, fig_all if create_plots else None)
        
        self._log_message("[bold green]所有省市数据处理完成！[/bold green]")
        
        # Complete summary visualizations
        results = self._calculate_principal_components_needed(singular_values_dict)
        
        if create_plots:
            self._finalize_summary_plots(fig_all, output_dirs, results)
        
        # Create summary Excel
        summary_excel_path = self._create_summary_excel(results, output_dirs)
        
        # Save singular values
        self._save_singular_values_to_excel(singular_values_dict, output_dirs['data_tables'])
        
        self._log_message("[bold green]分析完成！[/bold green]")
        
        return {
            'base_dir': output_dir,
            'output_dirs': output_dirs,
            'summary_excel': summary_excel_path,
            'singular_values_dict': singular_values_dict,
            'results': results
        }
    
    def _process_sheet(self, excel_file, sheet_name, singular_values_dict, output_dirs, create_plots, fig_all):
        """Process individual sheet for singular value analysis."""
        try:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            
            if df.shape[1] > 1:
                data = df.iloc[:, 1:].values.T  # Transpose to (years, indicators)
                data_centered = data - np.mean(data, axis=0)
                
                _, s, _ = svd(data_centered, full_matrices=False)
                singular_values_dict[sheet_name] = s
                
                explained_variance_ratio = s**2 / np.sum(s**2)
                cumulative_explained_variance = np.cumsum(explained_variance_ratio)
                
                if create_plots:
                    # Create individual province plot
                    self._create_province_plot(
                        sheet_name, s, explained_variance_ratio, output_dirs['single_plots']
                    )
                    
                    # Add to summary plot
                    if fig_all is not None:
                        fig_all.axes[0].plot(range(1, len(cumulative_explained_variance) + 1), 
                                           cumulative_explained_variance * 100, 'o-', 
                                           label=sheet_name, alpha=0.7)
        
        except Exception as e:
            self._log_message(f"[bold red]处理工作表 {sheet_name} 时出错: {e}[/bold red]")
    
    def _finalize_summary_plots(self, fig_all, output_dirs, results):
        """Complete and save summary visualization plots."""
        # Finalize all-provinces plot
        ax_all = fig_all.axes[0]
        ax_all.set_title('所有省市的累积方差贡献率', fontsize=16)
        ax_all.set_xlabel('主成分数量', fontsize=14)
        ax_all.set_ylabel('累积方差贡献率 (%)', fontsize=14)
        ax_all.grid(True, alpha=0.5)
        
        # Add threshold lines
        thresholds = [80, 85, 90, 95]
        for threshold, color in zip(thresholds, self.threshold_colors):
            ax_all.axhline(y=threshold, color=color, linestyle='--', 
                         label=f'{threshold}%阈值', alpha=0.7)
        
        ax_all.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize='small')
        plt.tight_layout()
        
        all_provinces_plot_path = os.path.join(output_dirs['summary_plots'], '所有省市累积方差贡献率.png')
        plt.savefig(all_provinces_plot_path, dpi=300, bbox_inches='tight')
        plt.close(fig_all)
        
        # Create average components plot
        df = pd.DataFrame(results)
        means = df[['80%方差所需主成分', '85%方差所需主成分', '90%方差所需主成分', '95%方差所需主成分']].mean()
        
        plt.figure(figsize=(10, 6))
        ax = means.plot(kind='bar', color=['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728'])
        plt.title('不同方差阈值下平均所需的主成分数量', fontsize=14)
        plt.xlabel('方差阈值', fontsize=12)
        plt.ylabel('平均主成分数量', fontsize=12)
        plt.grid(True, axis='y', alpha=0.5)
        
        for i, v in enumerate(means):
            ax.text(i, v + 0.1, f'{v:.1f}', ha='center', fontsize=10)
        
        plt.tight_layout()
        avg_pc_plot_path = os.path.join(output_dirs['summary_plots'], '平均主成分数量.png')
        plt.savefig(avg_pc_plot_path, dpi=300, bbox_inches='tight')
        plt.close()
        
        # Create distribution plots for each threshold
        threshold_columns = ['80%方差所需主成分', '85%方差所需主成分', '90%方差所需主成分', '95%方差所需主成分']
        thresholds = [0.8, 0.85, 0.9, 0.95]
        
        for threshold, column in zip(thresholds, threshold_columns):
            plt.figure(figsize=(15, 8))
            sorted_df = df.sort_values(by=column)
            ax = sorted_df.plot(kind='bar', x='省市', y=column, legend=False, color='#1f77b4', figsize=(15, 8))
            plt.title(f'{int(threshold*100)}%方差所需主成分数量 (按省市)', fontsize=14)
            plt.xlabel('省市', fontsize=12)
            plt.ylabel('主成分数量', fontsize=12)
            plt.xticks(rotation=90)
            plt.grid(True, axis='y', alpha=0.5)
            
            for i, v in enumerate(sorted_df[column]):
                ax.text(i, v + 0.1, str(v), ha='center', fontsize=9)
            
            plt.tight_layout()
            dist_plot_path = os.path.join(output_dirs['summary_plots'], f'{int(threshold*100)}%方差省市分布.png')
            plt.savefig(dist_plot_path, dpi=300, bbox_inches='tight')
            plt.close()
    
    def iterative_analysis(self, file_path, num_iterations=None, remove_percentage=None, output_dir=None):
        """
        Perform iterative PCA analysis with random indicator removal.
        
        Args:
            file_path: Path to Excel file
            num_iterations: Number of iterations (default: 20)
            remove_percentage: Percentage of indicators to remove (default: 0.5)
            output_dir: Output directory (auto-generated if None)
            
        Returns:
            dict: Summary of all iterations and final recommendations
        """
        if num_iterations is None:
            num_iterations = self.default_iterations
        if remove_percentage is None:
            remove_percentage = self.default_remove_percentage
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(file_path), '随机删除指标主成分分析')
        
        self._log_message(f"[bold yellow]开始执行 {num_iterations} 次随机删除指标的主成分分析...[/bold yellow]")
        
        all_results = []
        base_dir = output_dir
        
        # Iterative analysis
        for i in range(1, num_iterations + 1):
            self._log_message(f"\n[bold cyan]===== 开始第 {i} 次迭代 =====[/bold cyan]")
            
            # Randomly remove indicators
            temp_file, removed_indicators = self._randomly_remove_indicators(
                file_path, remove_percentage=remove_percentage
            )
            
            # Create iteration-specific output directory
            iteration_dirs = self._create_output_dirs(base_dir, iteration=i)
            
            # Save removed indicators list
            removed_file = os.path.join(iteration_dirs['data_tables'], '被删除的指标.xlsx')
            with pd.ExcelWriter(removed_file, engine='openpyxl') as writer:
                for sheet_name, indicators in removed_indicators.items():
                    pd.DataFrame({'被删除的指标': indicators}).to_excel(
                        writer, sheet_name=sheet_name, index=False)
            
            # Perform analysis
            iteration_result = self.analyze_singular_values(
                temp_file, 
                output_dir=iteration_dirs['root'],
                create_plots=True
            )
            
            iteration_result['iteration'] = i
            iteration_result['removed_indicators'] = removed_indicators
            all_results.append(iteration_result)
            
            time.sleep(1)  # Brief pause
        
        # Generate summary results
        self._log_message("\n[bold green]所有迭代完成，正在生成汇总分析...[/bold green]")
        summary_dir, summary_excel = self._summarize_all_iterations(all_results, base_dir)
        
        # Calculate final recommendations
        final_recommendations = self._calculate_final_recommendations(all_results)
        
        self._log_message("\n[bold green]=============== 分析完成 ===============[/bold green]")
        self._log_message(f"[yellow]汇总结果目录:[/yellow] {summary_dir}")
        self._log_message(f"[yellow]汇总数据Excel:[/yellow] {summary_excel}")
        
        # Display recommendations
        self._log_message("\n[bold cyan]主成分数量建议 (基于20次随机迭代):[/bold cyan]")
        for threshold, value in final_recommendations.items():
            self._log_message(f"[green]- {threshold}:[/green] {value:.1f}")
        
        return {
            'base_dir': base_dir,
            'summary_dir': summary_dir,
            'summary_excel': summary_excel,
            'all_results': all_results,
            'final_recommendations': final_recommendations
        }
    
    def _summarize_all_iterations(self, results_list, base_dir):
        """Summarize results from all iterations."""
        summary_dir = os.path.join(base_dir, '汇总结果')
        os.makedirs(summary_dir, exist_ok=True)
        
        # Extract iteration means
        iteration_means = []
        for i, result in enumerate(results_list):
            df = pd.DataFrame(result['results'])
            means = df[['80%方差所需主成分', '85%方差所需主成分', '90%方差所需主成分', '95%方差所需主成分']].mean()
            
            iteration_means.append({
                '迭代': i + 1,
                '80%方差平均主成分': means['80%方差所需主成分'],
                '85%方差平均主成分': means['85%方差所需主成分'],
                '90%方差平均主成分': means['90%方差所需主成分'],
                '95%方差平均主成分': means['95%方差所需主成分']
            })
        
        summary_df = pd.DataFrame(iteration_means)
        overall_means = summary_df[['80%方差平均主成分', '85%方差平均主成分', '90%方差平均主成分', '95%方差平均主成分']].mean()
        
        # Create summary plot
        plt.figure(figsize=(15, 8))
        
        for col, color in zip(
            ['80%方差平均主成分', '85%方差平均主成分', '90%方差平均主成分', '95%方差平均主成分'],
            ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728']
        ):
            plt.plot(summary_df['迭代'], summary_df[col], 'o-', label=col, color=color)
            plt.axhline(y=overall_means[col], color=color, linestyle='--', 
                       label=f'{col}总体平均: {overall_means[col]:.2f}', alpha=0.5)
        
        plt.title('20次随机迭代的主成分分析结果', fontsize=16)
        plt.xlabel('迭代次数', fontsize=14)
        plt.ylabel('平均主成分数量', fontsize=14)
        plt.grid(True, alpha=0.5)
        plt.legend(loc='upper right')
        plt.xticks(range(1, len(results_list) + 1))
        plt.tight_layout()
        plt.savefig(os.path.join(summary_dir, '20次迭代汇总结果.png'), dpi=300, bbox_inches='tight')
        plt.close()
        
        # Save summary Excel
        summary_excel = os.path.join(summary_dir, '20次迭代汇总数据.xlsx')
        with pd.ExcelWriter(summary_excel, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='各次迭代结果', index=False)
            
            pd.DataFrame({
                '统计项': ['总体平均值'],
                '80%方差平均主成分': [overall_means['80%方差平均主成分']],
                '85%方差平均主成分': [overall_means['85%方差平均主成分']],
                '90%方差平均主成分': [overall_means['90%方差平均主成分']],
                '95%方差平均主成分': [overall_means['95%方差平均主成分']]
            }).to_excel(writer, sheet_name='总体统计', index=False)
        
        return summary_dir, summary_excel
    
    def _calculate_final_recommendations(self, all_results):
        """Calculate final recommendations from all iterations."""
        iteration_means = []
        for result in all_results:
            df = pd.DataFrame(result['results'])
            means = df[['80%方差所需主成分', '85%方差所需主成分', '90%方差所需主成分', '95%方差所需主成分']].mean()
            iteration_means.append(means)
        
        summary_df = pd.DataFrame(iteration_means)
        overall_means = summary_df.mean()
        
        return {
            '80%方差平均所需主成分数': overall_means['80%方差所需主成分'],
            '85%方差平均所需主成分数': overall_means['85%方差所需主成分'],
            '90%方差平均所需主成分数': overall_means['90%方差所需主成分'],
            '95%方差平均所需主成分数': overall_means['95%方差所需主成分']
        }


def select_file():
    """Open file dialog to select Excel file."""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="选择Excel文件",
        filetypes=[("Excel文件", "*.xlsx;*.xls")]
    )
    root.destroy()
    return file_path


def create_gui():
    """Create a simple GUI for PCA analysis."""
    def run_single_analysis():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls")]
        )
        if file_path:
            try:
                analyzer = PCAAnalyzer()
                result = analyzer.analyze_singular_values(file_path)
                
                # Show completion message
                df = pd.DataFrame(result['results'])
                means = df[['80%方差所需主成分', '85%方差所需主成分', '90%方差所需主成分', '95%方差所需主成分']].mean()
                
                message = (
                    f"单次PCA分析完成！\n\n"
                    f"建议主成分数量:\n"
                    f"• 80%方差: {means['80%方差所需主成分']:.1f}\n"
                    f"• 85%方差: {means['85%方差所需主成分']:.1f}\n"
                    f"• 90%方差: {means['90%方差所需主成分']:.1f}\n"
                    f"• 95%方差: {means['95%方差所需主成分']:.1f}\n\n"
                    f"结果保存在:\n{result['base_dir']}"
                )
                
                messagebox.showinfo("分析完成", message)
                
            except Exception as e:
                messagebox.showerror("错误", f"分析失败:\n{str(e)}")
    
    def run_iterative_analysis():
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx;*.xls")]
        )
        if file_path:
            try:
                analyzer = PCAAnalyzer()
                result = analyzer.iterative_analysis(file_path)
                
                # Show completion message
                recommendations = result['final_recommendations']
                message = (
                    f"20次迭代PCA分析完成！\n\n"
                    f"最终建议主成分数量:\n"
                    f"• 80%方差: {recommendations['80%方差平均所需主成分数']:.1f}\n"
                    f"• 85%方差: {recommendations['85%方差平均所需主成分数']:.1f}\n"
                    f"• 90%方差: {recommendations['90%方差平均所需主成分数']:.1f}\n"
                    f"• 95%方差: {recommendations['95%方差平均所需主成分数']:.1f}\n\n"
                    f"结果保存在:\n{result['summary_dir']}"
                )
                
                messagebox.showinfo("分析完成", message)
                
            except Exception as e:
                messagebox.showerror("错误", f"分析失败:\n{str(e)}")
    
    # Create GUI
    root = tk.Tk()
    root.title("农业数据主成分分析工具")
    root.geometry("600x400")
    
    # Description
    description = tk.Label(
        root,
        text="农业数据主成分分析工具\n\n"
             "此工具提供两种分析模式：\n\n"
             "1. 单次分析：对完整数据进行PCA分析\n"
             "   • 计算每个省市的奇异值分布\n"
             "   • 分析不同方差阈值下的主成分需求\n"
             "   • 生成详细的可视化报告\n\n"
             "2. 迭代分析：随机删除50%指标，进行20次分析\n"
             "   • 测试模型的稳健性\n"
             "   • 提供更可靠的主成分数量建议\n"
             "   • 生成汇总统计和可视化\n\n"
             "请选择分析模式：",
        justify=tk.LEFT,
        padx=20,
        pady=20
    )
    description.pack(expand=True)
    
    # Buttons
    button_frame = tk.Frame(root)
    button_frame.pack(pady=20)
    
    single_button = tk.Button(
        button_frame,
        text="单次PCA分析",
        command=run_single_analysis,
        padx=20,
        pady=10,
        font=("Arial", 12),
        bg="#4CAF50",
        fg="white"
    )
    single_button.pack(side=tk.LEFT, padx=10)
    
    iterative_button = tk.Button(
        button_frame,
        text="迭代PCA分析 (20次)",
        command=run_iterative_analysis,
        padx=20,
        pady=10,
        font=("Arial", 12),
        bg="#2196F3",
        fg="white"
    )
    iterative_button.pack(side=tk.LEFT, padx=10)
    
    exit_button = tk.Button(
        button_frame,
        text="退出",
        command=root.destroy,
        padx=20,
        pady=10,
        font=("Arial", 12)
    )
    exit_button.pack(side=tk.LEFT, padx=10)
    
    root.mainloop()


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        # Command line usage
        input_file = sys.argv[1]
        mode = sys.argv[2] if len(sys.argv) > 2 else 'single'
        
        analyzer = PCAAnalyzer()
        
        if mode == 'iterative':
            result = analyzer.iterative_analysis(input_file)
            print(f"迭代分析完成。结果: {result['summary_excel']}")
        else:
            result = analyzer.analyze_singular_values(input_file)
            print(f"单次分析完成。结果: {result['summary_excel']}")
    else:
        # GUI usage
        try:
            create_gui()
        except ImportError:
            print("GUI需要tkinter库。使用基础版本...")
            analyzer = PCAAnalyzer()
            print("PCA分析器已准备就绪。请使用代码调用analyze_singular_values()或iterative_analysis()方法。")
