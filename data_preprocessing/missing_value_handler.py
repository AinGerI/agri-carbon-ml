#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Missing Value Handler for Agricultural Data

This module provides comprehensive missing value handling for Excel/CSV files using:
- Linear interpolation for single missing values
- Cubic spline interpolation for continuous missing sequences  
- Intelligent edge value estimation using trend analysis
- GUI interface for easy file selection and processing

Author: Thesis Research Project
Date: 2025
"""

import pandas as pd
import numpy as np
from scipy import interpolate
import os
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import csv


class MissingValueHandler:
    """
    A comprehensive missing value handler for agricultural data processing.
    
    Supports multiple file formats (CSV, XLS, XLSX) and provides various
    interpolation methods for different scenarios.
    """
    
    def __init__(self):
        """Initialize the handler with default province and year mappings."""
        # Define province mapping for standardization
        self.provinces = [
            "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林", "黑龙江", "上海", "江苏",
            "浙江", "安徽", "福建", "江西", "山东", "河南", "湖北", "湖南", "广东", "广西",
            "海南", "重庆", "四川", "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆"
        ]
        
        # Configurable year range - can be adjusted based on needs
        self.default_year_range = (2005, 2023)  # Default range
        self.extended_year_range = (2000, 2024)  # Extended range for compatibility
    
    def detect_year_range(self, df):
        """
        Automatically detect year range from DataFrame columns.
        
        Args:
            df: Input DataFrame
            
        Returns:
            tuple: (start_year, end_year) or None if no valid years found
        """
        years = []
        for col in df.columns[1:]:  # Skip first column (usually province names)
            try:
                year = int(str(col))
                if 1900 <= year <= 2100:
                    years.append(year)
            except (ValueError, TypeError):
                continue
        
        if years:
            return min(years), max(years)
        return None
    
    def fill_missing_values(self, file_path, year_range=None, auto_detect_years=True):
        """
        Read file and fill missing values using intelligent interpolation methods.
        
        Args:
            file_path: Path to the input file
            year_range: Tuple of (start_year, end_year). If None, uses default range
            auto_detect_years: If True, automatically detects year range from data
            
        Returns:
            dict: Dictionary of processed DataFrames, keyed by sheet name
        """
        # Determine file extension and processing method
        file_ext = os.path.splitext(file_path)[1].lower()
        
        try:
            processed_sheets = {}
            
            # Process CSV files
            if file_ext == '.csv':
                df_raw = pd.read_csv(file_path, header=None)
                
                if len(df_raw.index) > 1 and len(df_raw.columns) > 1:
                    data_values = df_raw.iloc[1:, 1:].values
                    
                    # Auto-detect years if enabled
                    if auto_detect_years:
                        temp_df = pd.DataFrame(df_raw.iloc[0, 1:].values).T
                        detected_range = self.detect_year_range(pd.concat([pd.DataFrame(['dummy']), temp_df], axis=1))
                        if detected_range:
                            year_range = detected_range
                    
                    # Use provided year range or default
                    if not year_range:
                        year_range = self.default_year_range
                    
                    years = [str(year) for year in range(year_range[0], year_range[1] + 1)]
                    
                    # Ensure data dimensions match
                    num_rows = min(len(data_values), len(self.provinces))
                    num_cols = min(data_values.shape[1], len(years))
                    
                    # Create DataFrame with proper indexing
                    index = self.provinces[:num_rows]
                    columns = years[:num_cols]
                    
                    df = pd.DataFrame(data_values[:num_rows, :num_cols], index=index, columns=columns)
                    df = self.process_dataframe(df)
                    
                    sheet_name = os.path.basename(file_path).split('.')[0]
                    processed_sheets[sheet_name] = df
                else:
                    raise Exception("CSV file format incorrect, needs at least 2 rows and 2 columns")
            
            # Process Excel files
            elif file_ext in ['.xls', '.xlsx']:
                # Use appropriate engine based on file type
                if file_ext == '.xls':
                    xl = pd.ExcelFile(file_path, engine='xlrd')
                else:
                    xl = pd.ExcelFile(file_path)
                
                # Process each worksheet
                for sheet_name in xl.sheet_names:
                    try:
                        df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                        
                        if len(df_raw.index) > 1 and len(df_raw.columns) > 1:
                            data_values = df_raw.iloc[1:, 1:].values
                            
                            # Auto-detect years if enabled
                            if auto_detect_years:
                                temp_df = pd.DataFrame(df_raw.iloc[0, 1:].values).T
                                detected_range = self.detect_year_range(pd.concat([pd.DataFrame(['dummy']), temp_df], axis=1))
                                if detected_range:
                                    year_range = detected_range
                            
                            # Use provided year range or default
                            if not year_range:
                                year_range = self.default_year_range
                            
                            years = [str(year) for year in range(year_range[0], year_range[1] + 1)]
                            
                            # Ensure data dimensions match
                            num_rows = min(len(data_values), len(self.provinces))
                            num_cols = min(data_values.shape[1], len(years))
                            
                            # Create DataFrame with proper indexing
                            index = self.provinces[:num_rows]
                            columns = years[:num_cols]
                            
                            df = pd.DataFrame(data_values[:num_rows, :num_cols], index=index, columns=columns)
                            df = self.process_dataframe(df)
                            
                            processed_sheets[sheet_name] = df
                        else:
                            print(f"Worksheet {sheet_name} has insufficient data, skipping")
                            continue
                            
                    except Exception as sheet_error:
                        print(f"Error processing worksheet {sheet_name}: {sheet_error}")
                        continue
            else:
                raise Exception(f"Unsupported file format: {file_ext}")
        
        except Exception as e:
            raise Exception(f"Unable to read or process file: {e}")
        
        return processed_sheets
    
    def process_dataframe(self, df):
        """
        Process DataFrame to fill missing values using intelligent interpolation.
        
        Args:
            df: Input DataFrame with potential missing values
            
        Returns:
            pd.DataFrame: DataFrame with missing values filled
        """
        # Identify missing values including Chinese spaces and other representations
        df = df.replace(['　', '', ' ', '-'], np.nan)
        
        # Convert all data to numeric type
        df = df.apply(pd.to_numeric, errors='coerce')
        
        # Process each row individually
        for idx, row in df.iterrows():
            filled_row = row.copy()
            
            # Get indices of non-null values
            valid_indices = np.where(~pd.isna(row))[0]
            
            if len(valid_indices) == 0:
                continue  # Skip if entire row is empty
            
            if len(valid_indices) == 1:
                # If only one non-null value, fill entire row with that value
                valid_value = row[valid_indices[0]]
                filled_row.fillna(valid_value, inplace=True)
                df.loc[idx] = filled_row
                continue
            
            # Use cubic spline interpolation if enough points available (≥4 points)
            if len(valid_indices) >= 4:
                known_x = valid_indices
                known_y = row[valid_indices].values
                
                # Create cubic spline interpolation function
                spline = interpolate.UnivariateSpline(known_x, known_y, k=3, s=0)
                
                # Fill missing values
                for col_idx in range(len(row)):
                    if pd.isna(row[col_idx]):
                        if min(valid_indices) <= col_idx <= max(valid_indices):
                            # Within range: use spline interpolation
                            filled_row[col_idx] = spline(col_idx)
                        else:
                            # Outside range: use trend-based extrapolation
                            filled_row[col_idx] = self._extrapolate_value(
                                row, valid_indices, col_idx
                            )
            else:
                # Use linear interpolation for fewer points
                filled_row = self._linear_interpolation_fill(row, valid_indices)
            
            # Update original DataFrame
            df.loc[idx] = filled_row
        
        return df
    
    def _extrapolate_value(self, row, valid_indices, col_idx):
        """
        Extrapolate values outside the range of known points using trend analysis.
        
        Args:
            row: Data row (pandas Series)
            valid_indices: Indices of valid (non-null) values
            col_idx: Index of the column to extrapolate
            
        Returns:
            float: Extrapolated value
        """
        if col_idx < min(valid_indices):
            # Forward extrapolation: based on first two non-null values
            if len(valid_indices) >= 2:
                first_val = row[valid_indices[0]]
                second_val = row[valid_indices[1]]
                trend = second_val - first_val
                distance = min(valid_indices) - col_idx
                weight = 1 / (distance + 1)
                return first_val - trend * distance * weight
            else:
                return row[min(valid_indices)]
        else:
            # Backward extrapolation: based on last two non-null values
            if len(valid_indices) >= 2:
                last_val = row[valid_indices[-1]]
                second_last_val = row[valid_indices[-2]]
                trend = last_val - second_last_val
                distance = col_idx - max(valid_indices)
                weight = 1 / (distance + 1)
                return last_val + trend * distance * weight
            else:
                return row[max(valid_indices)]
    
    def _linear_interpolation_fill(self, row, valid_indices):
        """
        Fill missing values using linear interpolation or averaging.
        
        Args:
            row: Data row (pandas Series)
            valid_indices: Indices of valid (non-null) values
            
        Returns:
            pd.Series: Row with missing values filled
        """
        filled_row = row.copy()
        
        for col_idx in range(len(row)):
            if pd.isna(row[col_idx]):
                # Find nearest valid values on left and right
                left_valid = valid_indices[valid_indices < col_idx] if any(valid_indices < col_idx) else None
                right_valid = valid_indices[valid_indices > col_idx] if any(valid_indices > col_idx) else None
                
                if left_valid is not None and right_valid is not None:
                    left_idx = max(left_valid)
                    right_idx = min(right_valid)
                    
                    # Direct average for adjacent values
                    if right_idx - left_idx == 2:
                        filled_row[col_idx] = (row[left_idx] + row[right_idx]) / 2
                    else:
                        # Linear interpolation for non-adjacent values
                        try:
                            known_x = valid_indices
                            known_y = row[valid_indices].values
                            f = interpolate.interp1d(known_x, known_y)
                            filled_row[col_idx] = f(col_idx)
                        except Exception:
                            # Fallback to simple average
                            filled_row[col_idx] = (row[left_idx] + row[right_idx]) / 2
                
                elif left_valid is not None or right_valid is not None:
                    # Edge cases: use trend-based estimation
                    filled_row[col_idx] = self._extrapolate_value(row, valid_indices, col_idx)
        
        return filled_row
    
    def apply_font_settings(self, workbook):
        """
        Apply font settings to Excel workbook for better readability.
        
        Args:
            workbook: Excel workbook object (openpyxl)
        """
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        value = str(cell.value)
                        
                        # Check if content contains digits or Chinese characters
                        contains_digits = bool(re.search(r'\d', value))
                        contains_chinese = bool(re.search(r'[\u4e00-\u9fff]', value))
                        
                        if contains_digits and not contains_chinese:
                            # Numbers: use Arial Unicode MS font
                            cell.font = Font(name='Arial Unicode MS')
                        elif contains_chinese:
                            # Chinese text: use Song font
                            cell.font = Font(name='宋体')
                        elif contains_digits and contains_chinese:
                            # Mixed: prioritize Chinese font
                            cell.font = Font(name='宋体')
    
    def process_files_in_folder(self, folder_path, output_folder=None, year_range=None):
        """
        Process all Excel and CSV files in a folder.
        
        Args:
            folder_path: Input folder path
            output_folder: Output folder path (default: creates "processed" subfolder)
            year_range: Tuple of (start_year, end_year) for processing
            
        Returns:
            int: Number of files processed
        """
        if output_folder is None:
            output_folder = os.path.join(folder_path, "processed")
        
        # Ensure output folder exists
        os.makedirs(output_folder, exist_ok=True)
        
        # Get all supported files
        supported_files = [f for f in os.listdir(folder_path) 
                          if f.lower().endswith(('.xls', '.xlsx', '.csv'))]
        
        if not supported_files:
            return 0
        
        processed_count = 0
        
        for file in supported_files:
            input_path = os.path.join(folder_path, file)
            file_name, file_ext = os.path.splitext(file)
            
            try:
                # Process file
                processed_sheets = self.fill_missing_values(input_path, year_range=year_range)
                
                # Generate output filename
                if file_ext.lower() == '.csv':
                    output_file = f"{file_name}_filled.xlsx"
                    output_path = os.path.join(output_folder, output_file)
                    
                    # Create new Excel file
                    wb = openpyxl.Workbook()
                    
                    # Remove default Sheet
                    if 'Sheet' in wb.sheetnames:
                        del wb['Sheet']
                    
                    # Write processed data to Excel file
                    for sheet_name, df in processed_sheets.items():
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # Write year headers in first row
                        for col_idx, year in enumerate(df.columns, start=2):
                            ws.cell(row=1, column=col_idx).value = year
                        
                        # Write province names in first column
                        for row_idx, province in enumerate(df.index, start=2):
                            ws.cell(row=row_idx, column=1).value = province
                        
                        # Write data
                        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                            for col_idx, value in enumerate(row, start=2):
                                ws.cell(row=row_idx, column=col_idx).value = value
                    
                    # Apply font settings
                    self.apply_font_settings(wb)
                    
                    # Save file
                    wb.save(output_path)
                else:
                    # For Excel files, create new file
                    output_file = f"{file_name}_filled{file_ext}"
                    output_path = os.path.join(output_folder, output_file)
                    
                    # Create new Excel file
                    wb = openpyxl.Workbook()
                    
                    # Remove default Sheet
                    if 'Sheet' in wb.sheetnames:
                        del wb['Sheet']
                    
                    # Write each processed worksheet
                    for sheet_name, df in processed_sheets.items():
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # Write year headers in first row
                        for col_idx, year in enumerate(df.columns, start=2):
                            ws.cell(row=1, column=col_idx).value = year
                        
                        # Write province names in first column
                        for row_idx, province in enumerate(df.index, start=2):
                            ws.cell(row=row_idx, column=1).value = province
                        
                        # Write data
                        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
                            for col_idx, value in enumerate(row, start=2):
                                ws.cell(row=row_idx, column=col_idx).value = value
                    
                    # Apply font settings
                    self.apply_font_settings(wb)
                    
                    # Save file
                    wb.save(output_path)
                
                processed_count += 1
                print(f"Processed: {file} -> {output_file}")
            except Exception as e:
                print(f"Error processing {file}: {e}")
        
        return processed_count


def simple_gui():
    """Provide a simple graphical user interface for the missing value handler."""
    
    # Create simple interface
    root = tk.Tk()
    root.title("Agricultural Data Missing Value Handler")
    root.geometry("500x350")
    
    handler = MissingValueHandler()
    
    # Select input folder
    def select_input():
        folder = filedialog.askdirectory(title="Select folder containing Excel files")
        if folder:
            # Process files in selected folder
            output_folder = os.path.join(folder, "processed")
            count = handler.process_files_in_folder(folder, output_folder)
            if count > 0:
                messagebox.showinfo("Success", f"Successfully processed {count} files.\nResults saved in: {output_folder}")
            else:
                messagebox.showinfo("Info", f"No Excel files found in the selected folder.")
    
    # Add description text
    label = tk.Label(root, text="Agricultural Data Missing Value Handler\n\n"
                                "Features:\n"
                                "- Supports CSV, XLS and XLSX formats\n"
                                "- Uses cubic spline interpolation for continuous missing values\n"
                                "- Processes all worksheets in Excel files\n"
                                "- Numbers use Arial Unicode MS font\n"
                                "- Chinese text uses Song font\n"
                                "- Auto-detects year range from data\n"
                                "- Intelligent trend-based extrapolation\n\n"
                                "Click the button below to select a folder to process.\n"
                                "Processed files will be saved to 'processed' subfolder.", 
                     padx=20, pady=20, justify=tk.LEFT)
    label.pack(expand=True)
    
    # Add button
    button = tk.Button(root, text="Select Folder", command=select_input, padx=20, pady=10)
    button.pack(pady=5)
    
    # Exit button
    exit_button = tk.Button(root, text="Exit", command=root.destroy, padx=10, pady=5)
    exit_button.pack(pady=5)
    
    root.mainloop()


if __name__ == "__main__":
    # Check command line arguments
    if len(sys.argv) > 1:
        # If command line arguments provided, process specified folder directly
        input_folder = sys.argv[1]
        output_folder = sys.argv[2] if len(sys.argv) > 2 else None
        
        print(f"Processing folder: {input_folder}")
        handler = MissingValueHandler()
        count = handler.process_files_in_folder(input_folder, output_folder)
        print(f"Successfully processed {count} files")
    else:
        # Otherwise launch the GUI
        simple_gui()